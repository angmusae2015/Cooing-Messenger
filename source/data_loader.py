from openpyxl import load_workbook
import pandas as pd
import re


wb = load_workbook('target/책배송-일지.xlsx', data_only=True)
book_info = "data/BookList.csv"
series_info = "data/SeriesList.csv"

book_data = pd.read_csv(book_info)
series_data = pd.read_csv(series_info)


class Series:   # class about series information
    def __init__(self, code):   # this requires code of the series
        self.code = None    # series code
        self.name = None    # series name
        self.step = None    # step of this series

        self.load_series_info(code)

    def load_series_info(self, code):     # load series info by code
        try:
            self.code, self.name, self.step = series_data.query(f'Code=="{code}"').iloc[0].to_list()

        except IndexError:
            raise self.SeriesNotFoundError

    class SeriesNotFoundError(Exception):
        def __str__(self):
            return "Could not found the given series"


class Book:  # class about book information
    def __init__(self, code, num=None, name=None):  # this requires code of series, and either number or a name.
        if (num or name) is None:
            raise self.NotEnoughInfoError  # raise this when both number and name is not given.

        self.series = Series(code)  # series of this book(Series)

        self.num = num  # number of this book(str)

        self.name = name  # name of this book(str)

        # self.load_book_info()

    """
    def load_book_info(self):   # load book info by series code and number or name


        raise self.BookNotFoundError    # raise this when book with given information is not found.
    """

    class NotEnoughInfoError(Exception):
        def __str__(self):
            return "Either name or number is required."

    class BookNotFoundError(Exception):
        def __str__(self):
            return "Could not found the given book."


class BookSchedule:
    def __init__(self, original_text):
        self.rex = {
            'series': re.compile('\w+'),  # finds series code (ex: JB, HR3)
            'step': re.compile('\[.+\]'),  # finds step (ex: [3], [2-1])
            'book list': re.compile('\(.+\)')  # finds book list (ex: (1~4), (1~2,5,7))
        }
        # self.step = None    # step of current series(float)
        self.book_list = []  # list of number of book([book])

        self.parse_text(original_text)

    def parse_text(self, original_text):  # parse schedule text
        for seriesSet in original_text.split('+'):  # split by series
            code = self.rex['series'].search(seriesSet)[0]

            """
            step_text = self.rex['step'].search(original_text)[0].strip('[]')  # parse step of series
            if step_text is not None:
                if len(step_text.split('-')) > 1:
                    self.step = tuple([int(i) for i in step_text.split('-')])
                else:
                    self.step = (int(step_text), 0)
            """

            book_list_text = self.rex['book list'].search(seriesSet)[0].strip('()')
            for book_range in book_list_text.split(','):
                if "~" in book_range:
                    start, end = [book_data.query(f'Series=="{code}"&Num=="{num}"').index[0] for num in
                                  book_range.split('~')]
                else:
                    start = book_data.query(f'Series=="{code}"&Num=="{book_range}"').index[0]
                    end = start

                for i in range(start, end + 1):
                    num, name = book_data.iloc[i]['Num':'Name'].to_list()
                    self.book_list.append(Book(code, num, name))


class DropDown:
    def __init__(self):
        self.menus = []  # list of menus([DropDown.Menu])
        self.selectedMenu = None  # selected menu(DropDown.Menu)

    def add_menus(self, title, value):  # add menus
        self.menus.append(self.Menu(title, value))

    def select(self):  # ask which menu user will select
        print("Type index")
        for i in range(0, len(self.menus)):
            print(f"{i + 1}. {self.menus[i].title}")

        index = int(input(">>  "))
        self.selectedMenu = self.menus[index - 1]
        print(f"You've selected {self.selectedMenu.title}")

    class Menu:  # Menu has two attributes: title(str), value
        def __init__(self, title, value):
            self.title = title  # title is a string that will be shown to user
            self.value = value  # value is a real data of menu


class Tab:
    def __init__(self):
        self.selected_sheet = None  # selected sheet(WorkSheet)
        self.selected_date = None  # selected date(MergedCellRange)
        self.schedule = {}  # schedules on selected date({child number(str) : book code(str)})

    def select_sheet(self):  # select sheet
        dd = DropDown()
        for sheet in wb.worksheets:
            dd.add_menus(sheet.title, sheet)
        dd.select()
        self.selected_sheet = dd.selectedMenu.value

    def get_departure_date(self):  # return every merged date cell's ranges(list of MergedCellRanges)
        departure_date_cells = []
        for mergedCell in self.selected_sheet.merged_cells.ranges:
            if mergedCell.start_cell.column == 1:
                departure_date_cells.append(mergedCell)

        return departure_date_cells

    def select_date(self):  # select merged date cell's range(MergedCellRange)
        departure_date_cells = self.get_departure_date()
        dd = DropDown()
        for dateCell in departure_date_cells:
            date = dateCell.start_cell
            dd.add_menus(f"{date.value.year}-{date.value.month}-{date.value.day}", dateCell)
        dd.select()
        self.selected_date = dd.selectedMenu.value

    def get_schedule(self):
        for i in range(self.selected_date.min_row, self.selected_date.max_row + 1):
            self.schedule[self.selected_sheet[f'D{i}'].value] = self.selected_sheet[f'H{i}'].value


if __name__ == '__main__':
    """
    t = Tab()
    t.select_sheet()
    t.select_date()
    t.get_schedule()
    """

    text = "RK[2](1-1~1-2,2-3~2-8)"
    sch = BookSchedule(text)
    for book in sch.book_list:
        print(book.name)
