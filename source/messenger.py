from openpyxl import load_workbook
import re
import csv


wb = load_workbook('target/책배송-일지.xlsx', data_only=True)
book_info = open("data/BookList.csv", 'r')
series_info = open("data/SeriesList.csv", 'r')


class Series:   # class about series information
    def __init__(self, code):   # this requires code of the series
        self.code = code    # series code
        self.name = None    # series name
        self.step = None    # step of this series

        self.load_series_info()

    def load_series_info(self):     # load series info by code
        for line in csv.reader(series_info):
            if line[0] == self.code:
                self.step = line[1]
                self.name = line[2]

                return 0

        raise self.SeriesNotFoundError  # raise this when series with given code is not found

    class SeriesNotFoundError(Exception):
        def __str__(self):
            return "Could not found the given series."


class Book:     # class about book information
    def __init__(self, code, num=None, name=None):  # this requires code of series, and either number or a name.
        if num or name is None:
            raise self.NotEnoughInfoError   # raise this when both number and name is not given.

        self.series = Series(code)  # series of this book
        self.num = num  # number of this book
        self.name = name    # name of this book

        self.load_book_info()

    def load_book_info(self):   # load book info by series code and number or name
        for line in csv.reader(book_info):
            if line[0] == self.series.code:
                self.series = Series(line[0])

                if self.num is not None:     # search book with number
                    if line[1] == self.num:
                        self.name = line[2]

                        return 0

                elif self.name is not None:  # search book with name
                    if line[2] == self.name:
                        self.num = line[1]

                        return 0

        raise self.BookNotFoundError    # raise this when book with given information is not found.

    class NotEnoughInfoError(Exception):
        def __str__(self):
            return "Either name or number is required."

    class BookNotFoundError(Exception):
        def __str__(self):
            return "Could not found the given book."


class BookSchedule:
    def __init__(self, original_text):
        self.rex = {
            'series': re.compile('\w+'),  # finds series code (ex: JB, HR3)
            'step': re.compile('\[.+\]'),  # finds step (ex: [3], [2-1])
            'book list': re.compile('\(.+\)')  # finds book list (ex: (1~4), (1~2,5,7))
        }
        self.original_text = original_text  # original text that is split by '+' from Excel file([str])
        # original text is like this: CH[3-2](1~2,5,8)

        self.step = None    # step of current series(float)
        self.book_list = []  # list of number of book([str])

    def parse_text(self):
        code = self.rex['series'].search(self.original_text)[0]

        step_text = self.rex['step'].search(self.original_text)[0].strip('[]')     # parse step of series
        if step_text is not None:
            self.step = float(step_text[0])
            if len(step_text) > 1:      # add half step(ex: 2-1 step is counted as 2.5 step)
                self.step = float(step_text[0]) + 0.5

        book_list_text = self.rex['book list'].search(self.original_text)[0].strip('()')    # parse list of books
        for book_range in book_list_text.split(','):
            start, end = book_range.split('~')


class DropDown:
    def __init__(self):
        self.menus = []  # list of menus([DropDown.Menu])
        self.selectedMenu = None  # selected menu(DropDown.Menu)

    def add_menus(self, title, value):  # add menus
        self.menus.append(self.Menu(title, value))

    def select(self):  # ask which menu user will select
        print("Type index")
        for i in range(0, len(self.menus)):
            print(f"{i + 1}. {self.menus[i].title}")

        index = int(input(">>  "))
        self.selectedMenu = self.menus[index - 1]
        print(f"You've selected {self.selectedMenu.title}")

    class Menu:  # Menu has two attributes: title(str), value
        def __init__(self, title, value):
            self.title = title  # title is a string that will be shown to user
            self.value = value  # value is a real data of menu


class Tab:
    def __init__(self):
        self.selected_sheet = None  # selected sheet(WorkSheet)
        self.selected_date = None  # selected date(MergedCellRange)
        self.schedule = {}  # schedules on selected date({child number(str) : book code(str)})

    def select_sheet(self):  # select sheet
        dd = DropDown()
        for sheet in wb.worksheets:
            dd.add_menus(sheet.title, sheet)
        dd.select()
        self.selected_sheet = dd.selectedMenu.value

    def get_departure_date(self):  # return every merged date cell's ranges(list of MergedCellRanges)
        departure_date_cells = []
        for mergedCell in self.selected_sheet.merged_cells.ranges:
            if mergedCell.start_cell.column == 1:
                departure_date_cells.append(mergedCell)

        return departure_date_cells

    def select_date(self):  # select merged date cell's range(MergedCellRange)
        departure_date_cells = self.get_departure_date()
        dd = DropDown()
        for dateCell in departure_date_cells:
            date = dateCell.start_cell
            dd.add_menus(f"{date.value.year}-{date.value.month}-{date.value.day}", dateCell)
        dd.select()
        self.selected_date = dd.selectedMenu.value

    def get_schedule(self):
        for i in range(self.selected_date.min_row, self.selected_date.max_row + 1):
            self.schedule[self.selected_sheet[f'D{i}'].value] = self.selected_sheet[f'H{i}'].value


t = Tab()
t.select_sheet()
t.select_date()
t.get_schedule()
