from PyQt5.QtWidgets import *
from openpyxl import load_workbook


class layout(QGridLayout):
    def __init__(self, window):
        super().__init__()

        self.window = window

        self.fileSelectLabel = QLabel('일지 파일 선택(.xlsx)')

        self.findButton = QPushButton("찾기", self.window)
        self.findButton.setFixedWidth(50)

        self.fileNameViewer = QTextBrowser()
        self.fileNameViewer.setFixedHeight(self.findButton.height() - 5)

        self.sheetSelectLabel = QLabel('불러올 시트 선택')

        self.sheetSelectBox = QComboBox()

        self.addWidget(self.fileSelectLabel, 0, 0, 1, 2)
        self.addWidget(self.findButton, 1, 0)
        self.addWidget(self.fileNameViewer, 1, 1, 1, 2)
        self.addWidget(self.sheetSelectLabel, 2, 0, 1, 2)
        self.addWidget(self.sheetSelectBox, 3, 0, 1, 3)

        self.findButton.clicked.connect(self.selectFile)

    def selectFile(self):
        fName = QFileDialog.getOpenFileName(self.window, 'Open file', 'C:/', 'Worksheet Files(*.xlsx)')
        self.window.selectedFile = load_workbook(fName[0])

        if fName[0]:
            self.fileNameViewer.setText(cutStr(fName[0]))

        self.addItemToComboBox()

    def addItemToComboBox(self):
        names = self.window.selectedFile.sheetnames

        for name in names:
            self.sheetSelectBox.addItem(name)


def cutStr(text):
    if len(text) > 28:
        fileName = text.split('/')[-1]
        text = text[:max(25 - len(fileName), 0)] + '.../' + fileName

    return text
